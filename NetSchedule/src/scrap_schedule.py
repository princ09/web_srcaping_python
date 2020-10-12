import json
import csv
from bs4 import BeautifulSoup
import requests
import re
import pymongo
import helper.DataProvider as dp
from datetime import timedelta, date
import xlsxwriter
from openpyxl import Workbook

today = date.today()
d1 = today.strftime("%d-%m-%Y")
mongoDBCon = pymongo.MongoClient("mongodb://localhost:27017/")

url = "https://wbes.wrldc.in/ReportNetSchedule/GetNetScheduleSummary"
file = '../public/net_schedule.xlsx'


# https://wbes.wrldc.in/ReportNetSchedule/GetNetScheduleRevisionNo?regionId=2&ScheduleDate=05-10-2020

def prepare_data(seller_or_buyer, date_pass, region, is_buyer):
    headers = dp.get_header()
    parameters = dp.get_schedule_params(seller_or_buyer, date_pass, region, is_buyer)
    print(parameters)
    response = requests.get(url, headers=headers, params=parameters)
    soup = BeautifulSoup(response.content, 'html.parser')
    data = soup.find_all('script')[2].string
    match = re.compile('var data = JSON.parse\((.*)\);')
    match = match.search(data.string)
    data = json.loads(match.group(1))
    data = data.replace('],[', ';')
    data = data.replace('[', '')
    data = data.replace('\"', '')
    print(data)
    return data


def get_schedule(seller_or_buyer, date_pass=d1, region="WEST", is_buyer=1):
    data = prepare_data(seller_or_buyer, date_pass, region, is_buyer)

    '''with open('data_store.csv', mode='w') as schedule_file:
        schedule_writer = csv.writer(schedule_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        for row in data.split(";"):
            temp = []
            for value in row.split(","):
                temp.append(value)
            schedule_writer.writerow(temp)'''
    with xlsxwriter.Workbook('data.xlsx') as workbook:
        worksheet = workbook.add_worksheet()
        for row_num, row in enumerate(data.split(";")):
            temp = []
            for value in row.split(","):
                temp.append(value)
            print(temp)
            worksheet.write_row(row_num, 0, temp)


# get_schedule('DNH_State', '09-10-2020', "WEST", 0)


def date_range(start_date_param, end_date_param):
    for n in range(int((end_date_param - start_date_param).days)):
        yield start_date_param + timedelta(n)


start_datee = date(2020, 6, 1)
end_datee = date.today() + timedelta(1)
book = Workbook()
sheet = book.active
for single_date in date_range(start_datee, end_datee):
    temp_date = single_date.strftime("%d-%m-%Y")
    data = prepare_data('DNH_State', temp_date, 'WEST', 0)
    for row_num, row in enumerate(data.split(";")):
        temp =[]
        for col, entry in enumerate(row.split(",")):
            if(row_num>0):
                if(col==1):
                    temp_date_value = temp_date+":"+entry.split("-")[0]
                    temp.append(temp_date_value)
                    continue
                elif (col==3):
                    continue
                else:
                    temp.append(entry)
            else:
                temp.append(entry)
            #ws.cell(row=row+row_num, column=col, value=entry)
        sheet.append(temp)

book.save("NetScheduleScrap.xlsx")